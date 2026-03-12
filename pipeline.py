"""
Pipeline de Auditoria Fiscal — Mercado de Curto Prazo (MCP)
============================================================

Contexto:
    Auditores fiscais da área de energia precisavam cruzar manualmente
    duas planilhas do sistema nacional da CCEE, empresa por empresa,
    para isolar os perfis de agentes com cargas em um estado específico
    e extrair apenas o Resultado Final proporcional àquelas cargas.

    Este pipeline automatiza esse processo em três etapas (ETL):
    1. Extração das planilhas de Consumo e Contabilização
    2. Transformação: identificar perfis com carga no estado alvo,
       calcular a proporção de carga por mês e aplicar sobre o
       Resultado Final da Contabilização
    3. Carga: exportar relatório Excel formatado para o auditor

AVISO LGPD:
    A versão de portfólio utiliza dados 100% fictícios gerados pelo
    script src/gerar_dados_simulados.py. Nenhum dado real de agentes,
    CNPJs ou valores financeiros foi utilizado ou exposto.

Autor: Austin
"""

import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
log = logging.getLogger(__name__)



# 1. EXTRACT

def extrair(caminho_consumo: str, caminho_contab: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    log.info("EXTRACT — lendo planilhas...")

    df_consumo = pd.read_excel(
        caminho_consumo,
        dtype={"CNPJ do Agente": str, "CNPJ da Carga": str, "CNPJ Agente - Uso do Fio": str}
    )
    df_contab = pd.read_excel(
        caminho_contab,
        dtype={"CNPJ Agente": str}
    )

    log.info(f"  CFZ004 Consumo:        {len(df_consumo):>6,} registros | "
             f"{df_consumo['Sigla do Perfil de Agente'].nunique()} perfis | "
             f"{df_consumo['UF da Carga'].nunique()} UFs")
    log.info(f"  CFZ003 Contabilização: {len(df_contab):>6,} registros | "
             f"{df_contab['Perfil de Agente'].nunique()} perfis")

    return df_consumo, df_contab



# 2. TRANSFORM

def transformar(df_consumo: pd.DataFrame, df_contab: pd.DataFrame, uf: str = "SE") -> pd.DataFrame:
    log.info(f"TRANSFORM — filtrando cargas em {uf} e calculando proporção...")

    # --- Padronização ---
    df_consumo["UF da Carga"] = df_consumo["UF da Carga"].str.strip().str.upper()
    df_consumo["Sigla do Perfil de Agente"] = df_consumo["Sigla do Perfil de Agente"].str.strip()
    df_consumo["CNPJ do Agente"] = df_consumo["CNPJ do Agente"].astype(str).str.strip()
    df_contab["Perfil de Agente"] = df_contab["Perfil de Agente"].str.strip()
    df_contab["CNPJ Agente"] = df_contab["CNPJ Agente"].astype(str).str.strip()

    # --- Passo 1: perfis com carga no estado alvo ---
    consumo_uf = df_consumo[df_consumo["UF da Carga"] == uf].copy()

    if consumo_uf.empty:
        log.warning(f"Nenhuma carga encontrada para UF={uf}.")
        return pd.DataFrame()

    perfis_uf = (
        consumo_uf[["CNPJ do Agente", "Sigla do Perfil de Agente"]]
        .drop_duplicates()
        .rename(columns={"CNPJ do Agente": "CNPJ Agente", "Sigla do Perfil de Agente": "Perfil de Agente"})
    )

    log.info(f"  Perfis com carga em {uf}: {len(perfis_uf)}")
    for _, r in perfis_uf.iterrows():
        log.info(f"    • {r['Perfil de Agente']}  (CNPJ {r['CNPJ Agente']})")

    # --- Passo 2: cruzamento com contabilização ---
    df_contab_filtrado = df_contab[[
        "Mês/Ano", "Evento", "Sigla do Agente", "CNPJ Agente",
        "Perfil de Agente", "Classe do Perfil",
        "Consumo Total - TRC a,s,r,w - (MWh)",
        "Resultado Final - RESULTADO a,m - (R$)"
    ]].copy()

    df_merged = perfis_uf.merge(df_contab_filtrado, on=["CNPJ Agente", "Perfil de Agente"], how="inner")

    if df_merged.empty:
        log.warning("Cruzamento não retornou registros. Verifique se os perfis existem na contabilização.")
        return pd.DataFrame()

    log.info(f"  Registros após cruzamento: {len(df_merged):,}")

    # --- Passo 3: carga total por perfil por mês ---
    carga_total = (
        df_consumo
        .groupby(["Ano/Mês", "Sigla do Perfil de Agente"])["Carga Medida - (MWh) "]
        .sum()
        .reset_index()
        .rename(columns={
            "Ano/Mês": "Mês/Ano",
            "Sigla do Perfil de Agente": "Perfil de Agente",
            "Carga Medida - (MWh) ": "Carga_Total_MWh"
        })
    )

    # --- Passo 4: carga do estado alvo por perfil por mês ---
    carga_uf = (
        consumo_uf
        .groupby(["Ano/Mês", "Sigla do Perfil de Agente"])
        .agg(
            Carga_UF_MWh=("Carga Medida - (MWh) ", "sum"),
            Pontos_Consumo_UF=("Pontos de Consumo", lambda x: " | ".join(sorted(x.dropna().unique()))),
            Cidades_UF=("Cidade da Carga", lambda x: ", ".join(sorted(x.dropna().unique())))
        )
        .reset_index()
        .rename(columns={
            "Ano/Mês": "Mês/Ano",
            "Sigla do Perfil de Agente": "Perfil de Agente"
        })
    )

    # --- Passo 5: proporção UF = carga_uf / carga_total por perfil/mês ---
    proporcao = carga_total.merge(carga_uf, on=["Mês/Ano", "Perfil de Agente"], how="inner")
    proporcao["Proporcao_UF"] = (proporcao["Carga_UF_MWh"] / proporcao["Carga_Total_MWh"]).round(6)

    # --- Passo 6: aplicar proporção sobre Resultado Final ---
    df_final = df_merged.merge(proporcao, on=["Mês/Ano", "Perfil de Agente"], how="left")

    df_final[f"Resultado_Proporcional_{uf}_RS"] = (
        df_final["Resultado Final - RESULTADO a,m - (R$)"] * df_final["Proporcao_UF"]
    ).round(2)

    df_final = df_final.sort_values(
        ["Sigla do Agente", "Perfil de Agente", "Mês/Ano"]
    ).reset_index(drop=True)

    log.info(f"  Proporção média de carga em {uf}: {df_final['Proporcao_UF'].mean():.2%}")

    resultado_nacional = df_final["Resultado Final - RESULTADO a,m - (R$)"].sum()
    resultado_uf = df_final[f"Resultado_Proporcional_{uf}_RS"].sum()
    log.info(f"  Resultado Final Nacional (perfis com carga em {uf}): R$ {resultado_nacional:>15,.2f}")
    log.info(f"  Resultado Final Proporcional {uf}:                   R$ {resultado_uf:>15,.2f}")

    return df_final



# 3. LOAD

def carregar(df: pd.DataFrame, df_consumo_raw: pd.DataFrame, caminho_saida: str, uf: str = "SE"):
    if df.empty:
        log.error("LOAD — DataFrame vazio. Nenhum arquivo gerado.")
        return

    log.info(f"LOAD — gerando relatório: {caminho_saida}")

    wb = Workbook()
    _aba_detalhe(wb, df, uf)
    _aba_resumo(wb, df, uf)
    _aba_rastreabilidade(wb, df_consumo_raw, df, uf)
    del wb["Sheet"]

    os.makedirs(os.path.dirname(caminho_saida), exist_ok=True)
    wb.save(caminho_saida)
    log.info("Relatório gerado com sucesso.")


# ── helpers de estilo ──────────────────────────
def _borda():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _cab(cell, cor="1F4E79"):
    cell.fill = PatternFill("solid", start_color=cor)
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _borda()


def _aba_detalhe(wb, df, uf):
    ws = wb.create_sheet(f"Detalhe_{uf}", 0)
    uf_col = [c for c in df.columns if c.startswith("Resultado_Proporcional_")][0]

    colunas = [
        "Mês/Ano", "Sigla do Agente", "CNPJ Agente", "Perfil de Agente", "Classe do Perfil",
        "Carga_Total_MWh", "Carga_UF_MWh", "Proporcao_UF",
        "Resultado Final - RESULTADO a,m - (R$)", uf_col,
        "Pontos_Consumo_UF", "Cidades_UF"
    ]

    for col, h in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _cab(cell)
    ws.row_dimensions[1].height = 42

    for r_idx, row in enumerate(df[colunas].itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.border = _borda()
            col_name = colunas[c_idx - 1]
            if "(R$)" in col_name:
                cell.number_format = "R$ #,##0.00;[RED](R$ #,##0.00)"
                if isinstance(val, (int, float)) and val < 0:
                    cell.font = Font(name="Arial", size=9, color="CC0000")
            elif "MWh" in col_name:
                cell.number_format = "#,##0.000"
            elif col_name == "Proporcao_UF":
                cell.number_format = "0.00%"

    larguras = [10, 16, 16, 26, 20, 16, 14, 12, 24, 24, 38, 22]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"


def _aba_resumo(wb, df, uf):
    ws = wb.create_sheet("Resumo_por_Perfil", 1)
    uf_col = [c for c in df.columns if c.startswith("Resultado_Proporcional_")][0]

    ws["A1"] = f"RESULTADO FINAL PROPORCIONAL — {uf} — REFERÊNCIA 2022"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79", name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:G1")

    ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  " \
               f"Dados: 100% fictícios — uso exclusivo de portfólio (LGPD)"
    ws["A2"].font = Font(italic=True, size=9, color="CC0000", name="Arial")
    ws.merge_cells("A2:G2")

    headers = ["Agente", "CNPJ", "Perfil de Agente", "Cidades UF",
               "Carga UF Total (MWh)", "Resultado Nacional (R$)", f"Resultado Proporcional {uf} (R$)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        _cab(cell, "375623")
    ws.row_dimensions[4].height = 35

    resumo = (
        df.groupby(["Sigla do Agente", "CNPJ Agente", "Perfil de Agente", "Cidades_UF"])
        .agg(
            Carga_UF=(       "Carga_UF_MWh",                            "sum"),
            Res_Nacional=(   "Resultado Final - RESULTADO a,m - (R$)",  "sum"),
            Res_Proporcional=(uf_col,                                    "sum"),
        )
        .reset_index()
        .sort_values("Res_Proporcional")
    )

    for r_idx, row in enumerate(resumo.itertuples(index=False), 5):
        vals = list(row)
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.border = _borda()
            col_name = headers[c_idx - 1]
            if "R$" in col_name:
                cell.number_format = "R$ #,##0.00;[RED](R$ #,##0.00)"
                if isinstance(val, (int, float)) and val < 0:
                    cell.font = Font(name="Arial", size=9, color="CC0000", bold=True)
            elif "MWh" in col_name:
                cell.number_format = "#,##0.000"

    # Linha de total
    total_row = len(resumo) + 5
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=10)
    for c_idx in range(1, 8):
        ws.cell(row=total_row, column=c_idx).border = _borda()
        ws.cell(row=total_row, column=c_idx).fill = PatternFill("solid", start_color="D9E1F2")

    ws.cell(row=total_row, column=5).value = f"=SUM(E5:E{total_row-1})"
    ws.cell(row=total_row, column=5).number_format = "#,##0.000"
    ws.cell(row=total_row, column=6).value = f"=SUM(F5:F{total_row-1})"
    ws.cell(row=total_row, column=6).number_format = "R$ #,##0.00;[RED](R$ #,##0.00)"
    ws.cell(row=total_row, column=7).value = f"=SUM(G5:G{total_row-1})"
    ws.cell(row=total_row, column=7).number_format = "R$ #,##0.00;[RED](R$ #,##0.00)"
    for c in [5, 6, 7]:
        ws.cell(row=total_row, column=c).font = Font(bold=True, name="Arial", size=10)

    larguras = [18, 16, 28, 30, 20, 26, 28]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"



def _aba_rastreabilidade(wb, df_consumo_raw, df_resultado, uf):
    ws = wb.create_sheet("Rastreabilidade_Evidencia", 2)

    ws["A1"] = f"EVIDENCIA DE RASTREABILIDADE — CARGAS EM {uf}"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79", name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:I1")

    ws["A2"] = (
        f"Esta aba comprova que apenas perfis com pontos de consumo em {uf} foram incluidos no resultado. "
        f"Cada linha e um ponto de consumo bruto da planilha CFZ004, filtrado por UF da Carga = '{uf}'."
    )
    ws["A2"].font = Font(italic=True, size=9, color="444444", name="Arial")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 30

    perfis_incluidos = set(df_resultado["Perfil de Agente"].unique())
    ws["A3"] = f"Perfis incluidos no resultado: {', '.join(sorted(perfis_incluidos))}"
    ws["A3"].font = Font(bold=True, size=9, color="375623", name="Arial")
    ws.merge_cells("A3:I3")
    ws.row_dimensions[3].height = 20

    df_consumo_raw = df_consumo_raw.copy()
    df_consumo_raw["UF da Carga"] = df_consumo_raw["UF da Carga"].str.strip().str.upper()
    df_uf = df_consumo_raw[df_consumo_raw["UF da Carga"] == uf].copy()
    df_uf = df_uf.sort_values(["Sigla do Perfil de Agente", "Ano/Mês", "Pontos de Consumo"]).reset_index(drop=True)

    colunas = [
        "Ano/Mês", "CNPJ do Agente", "Sigla do Perfil de Agente",
        "Pontos de Consumo", "Cidade da Carga", "UF da Carga",
        "Código do Ativo", "Carga Medida - (MWh) "
    ]
    headers = [
        "Mês", "CNPJ Agente", "Perfil de Agente",
        "Ponto de Consumo", "Cidade", "UF",
        "Código do Ativo", "Carga Medida (MWh)"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        _cab(cell, "7B2C2C")
    ws.row_dimensions[5].height = 30

    for r_idx, row in enumerate(df_uf[colunas].itertuples(index=False), 6):
        perfil_val = list(row)[2]
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.border = _borda()
            if headers[c_idx - 1] == "Carga Medida (MWh)":
                cell.number_format = "#,##0.000"
            if perfil_val in perfis_incluidos:
                cell.fill = PatternFill("solid", start_color="EBF1DE")

    larguras = [10, 18, 26, 30, 18, 6, 16, 18]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:{get_column_letter(len(headers))}5"

    log.info(f"  Rastreabilidade: {len(df_uf):,} pontos de consumo em {uf} documentados")


# EXECUÇÃO PRINCIPAL

def executar_pipeline(
    caminho_consumo: str,
    caminho_contab: str,
    caminho_saida: str,
    uf: str = "SE"
) -> pd.DataFrame:

    log.info("=" * 65)
    log.info("PIPELINE DE AUDITORIA FISCAL — MCP — CCEE")
    log.info("=" * 65)

    df_consumo, df_contab = extrair(caminho_consumo, caminho_contab)
    df_resultado = transformar(df_consumo, df_contab, uf=uf)
    carregar(df_resultado, df_consumo, caminho_saida, uf=uf)

    log.info("=" * 65)
    log.info("Pipeline concluído.")
    log.info("=" * 65)
    return df_resultado


if __name__ == "__main__":
    executar_pipeline(
        caminho_consumo="data/raw/CFZ004_Consumo_2022_SIMULADO.xlsx",
        caminho_contab="data/raw/CFZ003_Contabilizacao_2022_SIMULADO.xlsx",
        caminho_saida="output/relatorio_auditoria_SE_2022.xlsx",
        uf="SE"
    )
