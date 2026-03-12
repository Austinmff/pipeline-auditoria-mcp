"""
Gerador de Dados Simulados — Pipeline de Auditoria Fiscal MCP
==============================================================

AVISO LGPD / PRIVACIDADE:
    Este script gera dados INTEIRAMENTE FICTÍCIOS para fins de
    demonstração de portfólio. Nenhum dado real de agentes,
    CNPJs, valores financeiros ou qualquer informação proveniente
    de órgãos públicos ou sistemas nacionais foi utilizado.

    A estrutura das planilhas (nomes de colunas, formato de dados)
    foi reproduzida com base em dados públicos disponíveis no site
    da CCEE (Câmara de Comercialização de Energia Elétrica),
    conforme Lei de Acesso à Informação (Lei nº 12.527/2011).

Autor: Austin
"""

import random
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)


# Dados fictícios de agentes e perfis

AGENTES = [
    {
        "sigla": "DISTRIBEX",
        "cnpj": "12345678000190",
        "razao_social": "DISTRIBEX ENERGIA S/A",
        "perfis": [
            {"sigla": "DISTRIBEX NORDESTE", "classe": "Consumidor Livre", "ufs_principais": ["SE", "BA", "PE", "AL"]},
            {"sigla": "DISTRIBEX NORTE",    "classe": "Consumidor Livre", "ufs_principais": ["PA", "AM", "TO"]},
            {"sigla": "DISTRIBEX SUL",      "classe": "Consumidor Livre", "ufs_principais": ["PR", "SC", "RS"]},
            {"sigla": "DISTRIBEX I5 NE",    "classe": "Consumidor Especial", "ufs_principais": ["SE", "BA", "CE"]},
        ]
    },
    {
        "sigla": "INDUSPOWER",
        "cnpj": "98765432000155",
        "razao_social": "INDUSPOWER COMERCIALIZADORA LTDA",
        "perfis": [
            {"sigla": "INDUSPOWER SE/CO",   "classe": "Consumidor Livre",    "ufs_principais": ["GO", "MT", "MS", "DF"]},
            {"sigla": "INDUSPOWER NE",      "classe": "Consumidor Especial", "ufs_principais": ["CE", "RN", "PB", "SE"]},
            {"sigla": "INDUSPOWER SUL",     "classe": "Autoprodutor",        "ufs_principais": ["RS", "SC"]},
        ]
    },
    {
        "sigla": "ENERGITEC",
        "cnpj": "11223344000177",
        "razao_social": "ENERGITEC PARTICIPAÇÕES S/A",
        "perfis": [
            {"sigla": "ENERGITEC LIVRE NE", "classe": "Consumidor Livre",    "ufs_principais": ["SE", "BA", "AL", "PE"]},
            {"sigla": "ENERGITEC CONV",     "classe": "Consumidor Especial", "ufs_principais": ["SP", "RJ", "MG"]},
        ]
    },
]

PONTOS_POR_UF = {
    "SE": ["Planta Aracaju", "Unidade Lagarto", "Filial Itabaiana", "Dep. São Cristóvão"],
    "BA": ["Planta Salvador", "Unidade Feira", "Filial Vitória da Conquista"],
    "PE": ["Planta Recife", "Unidade Caruaru", "Filial Petrolina"],
    "AL": ["Planta Maceió", "Unidade Arapiraca"],
    "CE": ["Planta Fortaleza", "Unidade Juazeiro"],
    "PA": ["Planta Belém", "Unidade Santarém"],
    "AM": ["Planta Manaus"],
    "TO": ["Planta Palmas"],
    "PR": ["Planta Curitiba", "Unidade Londrina"],
    "SC": ["Planta Florianópolis", "Unidade Joinville"],
    "RS": ["Planta Porto Alegre", "Unidade Caxias do Sul"],
    "GO": ["Planta Goiânia", "Unidade Anápolis"],
    "MT": ["Planta Cuiabá"],
    "MS": ["Planta Campo Grande"],
    "DF": ["Planta Brasília"],
    "RN": ["Planta Natal"],
    "PB": ["Planta João Pessoa"],
    "SP": ["Planta São Paulo", "Unidade Campinas"],
    "RJ": ["Planta Rio de Janeiro"],
    "MG": ["Planta Belo Horizonte", "Unidade Uberlândia"],
}

DISTRIBUIDORAS_POR_UF = {
    "SE": ("ENERGISA SERGIPE - DISTRIBUIDORA DE ENERGIA S.A", "13017462000163"),
    "BA": ("COELBA - CIA ELETRICIDADE DO ESTADO DA BAHIA", "15139629000194"),
    "PE": ("CELPE - CIA ENERGETICA DE PERNAMBUCO", "10835932000108"),
    "AL": ("CEAL - CIA ENERGETICA DE ALAGOAS", "12272084000190"),
    "CE": ("ENEL DISTRIBUIÇÃO CEARÁ", "07047251000170"),
    "PA": ("CELPA - CENTRAIS ELETRICAS DO PARA S.A.", "04895728000180"),
    "AM": ("AMAZONAS ENERGIA S.A.", "02341467000120"),
    "TO": ("ENERGISA TOCANTINS", "25060933000141"),
    "PR": ("COPEL DISTRIBUIÇÃO S.A.", "04368703000166"),
    "SC": ("CELESC DISTRIBUIÇÃO S.A.", "08336783000190"),
    "RS": ("RGE SUL DISTRIBUIDORA DE ENERGIA S.A.", "02016440000162"),
    "GO": ("ENEL DISTRIBUIÇÃO GOIÁS", "01543032000132"),
    "MT": ("ENERGISA MATO GROSSO", "03940099000114"),
    "MS": ("ENERGISA MATO GROSSO DO SUL", "15413826000155"),
    "DF": ("CEB DISTRIBUIÇÃO S.A.", "08336783000190"),
    "RN": ("COSERN - CIA ENERGETICA DO RN", "08324196000105"),
    "PB": ("ENERGISA PARAÍBA", "09264856000176"),
    "SP": ("ENEL DISTRIBUIÇÃO SÃO PAULO", "10298355000190"),
    "RJ": ("LIGHT SERVIÇOS DE ELETRICIDADE S.A.", "03378521000175"),
    "MG": ("CEMIG DISTRIBUIÇÃO S.A.", "06981180000116"),
}

MESES = [f"2022/{str(m).zfill(2)}" for m in range(1, 13)]
EVENTO_TMPL = "{mes} - CONTABILIZAÇÃO"


def _borda():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _cabecalho(ws, headers, row=1, cor="1F4E79"):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = PatternFill("solid", start_color=cor)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _borda()
        ws.column_dimensions[get_column_letter(col)].width = max(len(h) * 1.1, 14)
    ws.row_dimensions[row].height = 35


def gerar_consumo(caminho: str):
    """
    CFZ004 — Planilha de Consumo
    Contém pontos de consumo por perfil de agente, com UF da carga.
    Estrutura fiel ao formato público da CCEE.
    Dados: 100% fictícios.
    """
    rng = random.Random(42)
    rows = []

    for agente in AGENTES:
        for perfil in agente["perfis"]:
            ufs = perfil["ufs_principais"]
            # Gera pontos de consumo por UF — cada perfil tem mais pontos nas UFs principais
            for uf in ufs:
                n_pontos = rng.randint(1, 4)
                pontos_disponiveis = PONTOS_POR_UF.get(uf, [f"Unidade {uf}"])
                pontos = rng.choices(pontos_disponiveis, k=min(n_pontos, len(pontos_disponiveis)))
                dist_razao, dist_cnpj = DISTRIBUIDORAS_POR_UF.get(uf, ("DISTRIBUIDORA LOCAL", "00000000000100"))

                for ponto in set(pontos):
                    carga_base = rng.uniform(200, 15000)
                    for mes in MESES:
                        sazonalidade = 1 + rng.uniform(-0.15, 0.20)
                        carga = round(carga_base * sazonalidade, 6)
                        rows.append({
                            "Ano/Mês": mes,
                            "Evento": EVENTO_TMPL.format(mes=mes),
                            "CNPJ do Agente": agente["cnpj"],
                            "Razão Social do Agente Proprietário": agente["razao_social"],
                            "Sigla do Perfil de Agente": perfil["sigla"],
                            "Classe do Agente Proprietário": perfil["classe"],
                            "Código da Parcela de Ativo": rng.randint(100000, 999999),
                            "Pontos de Consumo": ponto,
                            "Percentual de Propriedade (%)": "100%",
                            "CNPJ da Carga": agente["cnpj"][:8] + str(rng.randint(10000, 99999)),
                            "Endereço da Carga": f"RUA FICTÍCIA, {rng.randint(1, 9999)}",
                            "Bairro da Carga": "CENTRO",
                            "Cidade da Carga": PONTOS_POR_UF.get(uf, ["Capital"])[0].replace("Planta ", "").replace("Unidade ", "").replace("Filial ", "").replace("Dep. ", ""),
                            "UF da Carga": uf,
                            "Código do Ativo": f"{uf}{str(rng.randint(100000, 999999)).zfill(6)}",
                            "Razão Social - Uso do Fio": dist_razao,
                            "CNPJ Agente - Uso do Fio": dist_cnpj,
                            "Participação (%)": round(rng.uniform(0.01, 1.0), 6),
                            "Carga Medida - (MWh) ": carga,
                        })

    df = pd.DataFrame(rows).sort_values(["Ano/Mês", "Sigla do Perfil de Agente"]).reset_index(drop=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = list(df.columns)
    _cabecalho(ws, headers)

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.border = _borda()
            if headers[c_idx - 1] == "Carga Medida - (MWh) ":
                cell.number_format = "#,##0.000000"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wb.save(caminho)
    print(f"[OK] Consumo gerado: {caminho} ({len(df):,} registros)")
    return df


def gerar_contabilizacao(caminho: str):
    """
    CFZ003 — Planilha de Contabilização
    Contém resultado financeiro mensal por perfil de agente no MCP.
    Estrutura fiel ao formato público da CCEE.
    Dados: 100% fictícios.
    """
    rng = random.Random(99)
    rows = []

    for agente in AGENTES:
        for perfil in agente["perfis"]:
            consumo_base = rng.uniform(1000, 50000)
            resultado_base = rng.uniform(-2000000, 500000)

            for mes in MESES:
                sazon = 1 + rng.uniform(-0.20, 0.25)
                consumo = round(consumo_base * sazon, 6)
                resultado = round(resultado_base * sazon, 2)
                result_preliminar = round(resultado * rng.uniform(0.95, 1.05), 2)
                ajustes = round(resultado - result_preliminar, 2)
                tm_mcp = round(resultado * rng.uniform(0.40, 0.70), 2)
                encargos = round(abs(resultado) * rng.uniform(0.05, 0.15) * -1, 2)

                rows.append({
                    "Mês/Ano": mes,
                    "Evento": EVENTO_TMPL.format(mes=mes),
                    "Sigla do Agente": agente["sigla"],
                    "CNPJ Agente": agente["cnpj"],
                    "Perfil de Agente": perfil["sigla"],
                    "Classe do Perfil": perfil["classe"],
                    "Garantia Física Total - TGFIS a,w,r - (MWh)": None,
                    "Geração Total - TGG a,s,r,w - (MWh)": None,
                    "Consolidação do Resultado do MRE - MRE a,s,r,w - (MWh)": None,
                    "Consumo da Geração - TGGC a,s,r,w - (MWh)": None,
                    "Consumo Total - TRC a,s,r,w - (MWh)": consumo,
                    "Contratação Líquida - PCL a,s,w,r - (MWh)": round(consumo * rng.uniform(-0.05, 0.05), 6),
                    "Contratos de Venda Total - (MWh)": 0,
                    "Contratos de Compra Total - (MWh)": round(consumo * rng.uniform(0.95, 1.05), 6),
                    "Compensação do MRE - COMPENSAÇÃO_MRE a,m - (R$)": None,
                    "Total Mensal do Resultado no Mercado de Curto Prazo - TM_MCP a,m - (R$)": tm_mcp,
                    "Total de Ajustes de Exposições Financeiras - TAJ_EF a,m - (R$)": 0,
                    "Total de Encargos Consolidados - ENCARGOS a,m - (R$)": encargos,
                    "Total de Ajustes referente ao Alívio Retroativo - TAJ_AR a,m - (R$)": 0.0,
                    "Resultado da Exportação do MRE - RES_EXP_MRE a,m - (R$)": None,
                    "Efeito da Contratação de Itaipu - EC_IT a,m - (R$)": 0,
                    "Efeito de Repasse do Risco Hidrológico - ERRH a,m - (R$)": 0,
                    "Efeito da Contratação por Disponibilidade - ECD a,m - (R$)": None,
                    "Efeito da Contratação de Cotas de Garantia Física - ECCGFa,m - (R$)": 0,
                    "Efeito da Contratação de Comercialização de Energia Nuclear- ECCEN a,m (R$)": 0,
                    "Ajuste Decorrente de Recontabilizações - AJU_RECON a,m - (R$)": round(rng.uniform(-50000, 50000), 2),
                    "Ajuste Decorrente de Simulações - AJU_SIM a,m - (R$)": round(rng.uniform(-5000, 5000), 2),
                    "Resultado Referente ao Excedente Financeiro da Energia de Reserva - RES_EXCD_ER a,m (R$)": 0,
                    "Ajuste Decorrente do MCSD Ex-Post - MCSD_XP a,m - (R$)": None,
                    "Efeito dos Custos devido ao descolamento entre PLD e CMO - E_DESC a,m - (R$)": round(rng.uniform(-10000, 0), 2),
                    "Resultado Preliminar - RES_PRE a,m - (R$)": result_preliminar,
                    "Resultado Final - RESULTADO a,m - (R$)": resultado,
                    "Ajustes - (R$)": ajustes,
                    "Valor a liquidar pelo perfil de agente - (R$)": round(resultado * rng.uniform(1.8, 2.2), 2),
                })

    df = pd.DataFrame(rows).sort_values(["Mês/Ano", "Sigla do Agente"]).reset_index(drop=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = list(df.columns)
    _cabecalho(ws, headers)

    moeda_cols = [h for h in headers if "(R$)" in h]
    mwh_cols = [h for h in headers if "(MWh)" in h]

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.border = _borda()
            col_name = headers[c_idx - 1]
            if col_name in moeda_cols and val is not None:
                cell.number_format = "R$ #,##0.00;[RED](R$ #,##0.00)"
            elif col_name in mwh_cols and val is not None:
                cell.number_format = "#,##0.000000"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wb.save(caminho)
    print(f"[OK] Contabilização gerada: {caminho} ({len(df):,} registros)")
    return df


if __name__ == "__main__":
    gerar_consumo("data/raw/CFZ004_Consumo_2022_SIMULADO.xlsx")
    gerar_contabilizacao("data/raw/CFZ003_Contabilizacao_2022_SIMULADO.xlsx")
    print("\n[AVISO] Todos os dados gerados são 100% fictícios.")
    print("[AVISO] Nenhum dado real foi utilizado (LGPD — Lei 13.709/2018).")
